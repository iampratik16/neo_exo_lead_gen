"""
Neo Eco Cleaning — Google Maps Lead Scraper Engine
====================================================
Two-pass architecture (inherited from the original repo):
  Pass 1: Scroll Google Maps feed, collect all business metadata.
  Pass 2: Visit each website in a fresh page to extract emails.

Adapted for Neo Eco Cleaning Services Ltd — targeting property management
firms and estate/letting agents in and around London (primarily North London).

CRITICAL: Uses geolocation spoofing + Maps viewport coordinates to ensure
results are centred on London, not the user's physical location.

v2.1 — Added dynamic query filtering, AI-email enrichment fields,
       and CSV export alongside Excel.
"""

import asyncio
import csv
import os
import re
import urllib.parse
from dataclasses import dataclass, field
from datetime import datetime
from typing import Callable, Dict, List, Optional, Set, Tuple

from bs4 import BeautifulSoup
from playwright.async_api import async_playwright, Page, BrowserContext

from models import SearchRequest, Lead, ScrapeProgress


# ---------------------------------------------------------------------------
# London coordinates for geolocation spoofing
# ---------------------------------------------------------------------------
LONDON_LAT = 51.5074
LONDON_LNG = -0.1278
LONDON_ZOOM = 12


# ---------------------------------------------------------------------------
# Search Queries — exactly as specified in master.md
# Each tuple: (search_term, location, icp_tier, category, business_type_tag)
# The business_type_tag links each query to the frontend filter chips.
# ---------------------------------------------------------------------------
ICP_TIER_1_LABEL = "Tier 1 \u2013 Block Management"
ICP_TIER_2_LABEL = "Tier 2 \u2013 Estate Agent"

CATEGORY_PROPERTY_MGMT = "Property Management"
CATEGORY_ESTATE_AGENT = "Estate Agent / Letting Agent"

# business_type_tag values (must match the filter chip IDs in /api/filters)
BT_BLOCK_MGMT = "block_management"
BT_PROPERTY_MGMT = "property_management"
BT_LEASEHOLD = "leasehold_management"
BT_RTM = "rtm_management"
BT_SERVICE_CHARGE = "service_charge"
BT_ESTATE_AGENT = "estate_agent"
BT_LETTING_AGENT = "letting_agent"

SEARCH_QUERIES: List[Tuple[str, str, str, str, str]] = [
    # ---- ICP 1: Property Management (Tier 1 targets) ----
    ("block management company", "Barnet, London", ICP_TIER_1_LABEL, CATEGORY_PROPERTY_MGMT, BT_BLOCK_MGMT),
    ("block management company", "Enfield, London", ICP_TIER_1_LABEL, CATEGORY_PROPERTY_MGMT, BT_BLOCK_MGMT),
    ("block management company", "Haringey, London", ICP_TIER_1_LABEL, CATEGORY_PROPERTY_MGMT, BT_BLOCK_MGMT),
    ("block management company", "Islington, London", ICP_TIER_1_LABEL, CATEGORY_PROPERTY_MGMT, BT_BLOCK_MGMT),
    ("block management company", "Camden, London", ICP_TIER_1_LABEL, CATEGORY_PROPERTY_MGMT, BT_BLOCK_MGMT),
    ("block management company", "Finchley, London", ICP_TIER_1_LABEL, CATEGORY_PROPERTY_MGMT, BT_BLOCK_MGMT),
    ("block management company", "Muswell Hill, London", ICP_TIER_1_LABEL, CATEGORY_PROPERTY_MGMT, BT_BLOCK_MGMT),
    ("block management company", "Highgate, London", ICP_TIER_1_LABEL, CATEGORY_PROPERTY_MGMT, BT_BLOCK_MGMT),
    ("block management company", "Crouch End, London", ICP_TIER_1_LABEL, CATEGORY_PROPERTY_MGMT, BT_BLOCK_MGMT),
    ("block management company", "Tottenham, London", ICP_TIER_1_LABEL, CATEGORY_PROPERTY_MGMT, BT_BLOCK_MGMT),
    ("block management company", "Wood Green, London", ICP_TIER_1_LABEL, CATEGORY_PROPERTY_MGMT, BT_BLOCK_MGMT),
    ("block management company", "Archway, London", ICP_TIER_1_LABEL, CATEGORY_PROPERTY_MGMT, BT_BLOCK_MGMT),
    ("block management company", "Kentish Town, London", ICP_TIER_1_LABEL, CATEGORY_PROPERTY_MGMT, BT_BLOCK_MGMT),
    ("block management company", "Hampstead, London", ICP_TIER_1_LABEL, CATEGORY_PROPERTY_MGMT, BT_BLOCK_MGMT),
    ("block management company", "Finsbury Park, London", ICP_TIER_1_LABEL, CATEGORY_PROPERTY_MGMT, BT_BLOCK_MGMT),
    ("block management company", "Holloway, London", ICP_TIER_1_LABEL, CATEGORY_PROPERTY_MGMT, BT_BLOCK_MGMT),
    ("block management company", "Angel, London", ICP_TIER_1_LABEL, CATEGORY_PROPERTY_MGMT, BT_BLOCK_MGMT),
    ("block management company", "Palmers Green, London", ICP_TIER_1_LABEL, CATEGORY_PROPERTY_MGMT, BT_BLOCK_MGMT),
    ("block management company", "Edgware, London", ICP_TIER_1_LABEL, CATEGORY_PROPERTY_MGMT, BT_BLOCK_MGMT),
    ("property management company", "Edgware, London", ICP_TIER_1_LABEL, CATEGORY_PROPERTY_MGMT, BT_PROPERTY_MGMT),
    # Harrow & surrounding areas
    ("block management company", "Harrow, London", ICP_TIER_1_LABEL, CATEGORY_PROPERTY_MGMT, BT_BLOCK_MGMT),
    ("block management company", "Pinner, London", ICP_TIER_1_LABEL, CATEGORY_PROPERTY_MGMT, BT_BLOCK_MGMT),
    ("block management company", "Stanmore, London", ICP_TIER_1_LABEL, CATEGORY_PROPERTY_MGMT, BT_BLOCK_MGMT),
    ("block management company", "Northwood, London", ICP_TIER_1_LABEL, CATEGORY_PROPERTY_MGMT, BT_BLOCK_MGMT),
    ("block management company", "Ruislip, London", ICP_TIER_1_LABEL, CATEGORY_PROPERTY_MGMT, BT_BLOCK_MGMT),
    ("block management company", "Eastcote, London", ICP_TIER_1_LABEL, CATEGORY_PROPERTY_MGMT, BT_BLOCK_MGMT),
    ("block management company", "Kingsbury, London", ICP_TIER_1_LABEL, CATEGORY_PROPERTY_MGMT, BT_BLOCK_MGMT),
    ("property management company", "Harrow, London", ICP_TIER_1_LABEL, CATEGORY_PROPERTY_MGMT, BT_PROPERTY_MGMT),
    ("property management company", "North London", ICP_TIER_1_LABEL, CATEGORY_PROPERTY_MGMT, BT_PROPERTY_MGMT),
    ("leasehold management company", "London", ICP_TIER_1_LABEL, CATEGORY_PROPERTY_MGMT, BT_LEASEHOLD),
    ("residential block management", "London", ICP_TIER_1_LABEL, CATEGORY_PROPERTY_MGMT, BT_BLOCK_MGMT),
    ("RTM management company", "London", ICP_TIER_1_LABEL, CATEGORY_PROPERTY_MGMT, BT_RTM),
    ("service charge management", "London", ICP_TIER_1_LABEL, CATEGORY_PROPERTY_MGMT, BT_SERVICE_CHARGE),

    # ---- ICP 2: Estate & Letting Agents (Tier 2 targets) ----
    ("estate agent", "Barnet, London", ICP_TIER_2_LABEL, CATEGORY_ESTATE_AGENT, BT_ESTATE_AGENT),
    ("letting agent", "Barnet, London", ICP_TIER_2_LABEL, CATEGORY_ESTATE_AGENT, BT_LETTING_AGENT),
    ("estate agent", "Enfield, London", ICP_TIER_2_LABEL, CATEGORY_ESTATE_AGENT, BT_ESTATE_AGENT),
    ("letting agent", "Enfield, London", ICP_TIER_2_LABEL, CATEGORY_ESTATE_AGENT, BT_LETTING_AGENT),
    ("estate agent", "Haringey, London", ICP_TIER_2_LABEL, CATEGORY_ESTATE_AGENT, BT_ESTATE_AGENT),
    ("letting agent", "Haringey, London", ICP_TIER_2_LABEL, CATEGORY_ESTATE_AGENT, BT_LETTING_AGENT),
    ("estate agent", "Islington, London", ICP_TIER_2_LABEL, CATEGORY_ESTATE_AGENT, BT_ESTATE_AGENT),
    ("letting agent", "Islington, London", ICP_TIER_2_LABEL, CATEGORY_ESTATE_AGENT, BT_LETTING_AGENT),
    ("estate agent", "Camden, London", ICP_TIER_2_LABEL, CATEGORY_ESTATE_AGENT, BT_ESTATE_AGENT),
    ("letting agent", "Camden, London", ICP_TIER_2_LABEL, CATEGORY_ESTATE_AGENT, BT_LETTING_AGENT),
    ("estate agent", "Muswell Hill, London", ICP_TIER_2_LABEL, CATEGORY_ESTATE_AGENT, BT_ESTATE_AGENT),
    ("estate agent", "Highgate, London", ICP_TIER_2_LABEL, CATEGORY_ESTATE_AGENT, BT_ESTATE_AGENT),
    ("estate agent", "Crouch End, London", ICP_TIER_2_LABEL, CATEGORY_ESTATE_AGENT, BT_ESTATE_AGENT),
    ("estate agent", "Finchley, London", ICP_TIER_2_LABEL, CATEGORY_ESTATE_AGENT, BT_ESTATE_AGENT),
    ("estate agent", "Tottenham, London", ICP_TIER_2_LABEL, CATEGORY_ESTATE_AGENT, BT_ESTATE_AGENT),
    ("estate agent", "Wood Green, London", ICP_TIER_2_LABEL, CATEGORY_ESTATE_AGENT, BT_ESTATE_AGENT),
    ("estate agent", "Archway, London", ICP_TIER_2_LABEL, CATEGORY_ESTATE_AGENT, BT_ESTATE_AGENT),
    ("estate agent", "Kentish Town, London", ICP_TIER_2_LABEL, CATEGORY_ESTATE_AGENT, BT_ESTATE_AGENT),
    ("estate agent", "Hampstead, London", ICP_TIER_2_LABEL, CATEGORY_ESTATE_AGENT, BT_ESTATE_AGENT),
    ("estate agent", "Finsbury Park, London", ICP_TIER_2_LABEL, CATEGORY_ESTATE_AGENT, BT_ESTATE_AGENT),
    ("estate agent", "Holloway, London", ICP_TIER_2_LABEL, CATEGORY_ESTATE_AGENT, BT_ESTATE_AGENT),
    ("estate agent", "Angel, London", ICP_TIER_2_LABEL, CATEGORY_ESTATE_AGENT, BT_ESTATE_AGENT),
    ("estate agent", "Palmers Green, London", ICP_TIER_2_LABEL, CATEGORY_ESTATE_AGENT, BT_ESTATE_AGENT),
    ("estate agent", "Edgware, London", ICP_TIER_2_LABEL, CATEGORY_ESTATE_AGENT, BT_ESTATE_AGENT),
    ("letting agent", "Edgware, London", ICP_TIER_2_LABEL, CATEGORY_ESTATE_AGENT, BT_LETTING_AGENT),
    ("letting agent", "North London", ICP_TIER_2_LABEL, CATEGORY_ESTATE_AGENT, BT_LETTING_AGENT),
    # Harrow & surrounding areas
    ("estate agent", "Harrow, London", ICP_TIER_2_LABEL, CATEGORY_ESTATE_AGENT, BT_ESTATE_AGENT),
    ("letting agent", "Harrow, London", ICP_TIER_2_LABEL, CATEGORY_ESTATE_AGENT, BT_LETTING_AGENT),
    ("estate agent", "Pinner, London", ICP_TIER_2_LABEL, CATEGORY_ESTATE_AGENT, BT_ESTATE_AGENT),
    ("letting agent", "Pinner, London", ICP_TIER_2_LABEL, CATEGORY_ESTATE_AGENT, BT_LETTING_AGENT),
    ("estate agent", "Stanmore, London", ICP_TIER_2_LABEL, CATEGORY_ESTATE_AGENT, BT_ESTATE_AGENT),
    ("estate agent", "Northwood, London", ICP_TIER_2_LABEL, CATEGORY_ESTATE_AGENT, BT_ESTATE_AGENT),
    ("estate agent", "Ruislip, London", ICP_TIER_2_LABEL, CATEGORY_ESTATE_AGENT, BT_ESTATE_AGENT),
    ("estate agent", "Eastcote, London", ICP_TIER_2_LABEL, CATEGORY_ESTATE_AGENT, BT_ESTATE_AGENT),
    ("estate agent", "Kingsbury, London", ICP_TIER_2_LABEL, CATEGORY_ESTATE_AGENT, BT_ESTATE_AGENT),
]


# ---------------------------------------------------------------------------
# Available filter options (served by /api/filters)
# ---------------------------------------------------------------------------
AVAILABLE_BUSINESS_TYPES = [
    {"id": BT_BLOCK_MGMT, "label": "Block Management Companies", "tier": "Tier 1"},
    {"id": BT_PROPERTY_MGMT, "label": "Property Management Companies", "tier": "Tier 1"},
    {"id": BT_LEASEHOLD, "label": "Leasehold Management", "tier": "Tier 1"},
    {"id": BT_RTM, "label": "RTM Management Companies", "tier": "Tier 1"},
    {"id": BT_SERVICE_CHARGE, "label": "Service Charge Management", "tier": "Tier 1"},
    {"id": BT_ESTATE_AGENT, "label": "Estate Agents", "tier": "Tier 2"},
    {"id": BT_LETTING_AGENT, "label": "Letting Agents", "tier": "Tier 2"},
]

AVAILABLE_LOCATIONS = [
    {"id": "all", "label": "All London & Surrounding", "group": "Quick"},
    {"id": "Barnet, London", "label": "Barnet", "group": "North London"},
    {"id": "Enfield, London", "label": "Enfield", "group": "North London"},
    {"id": "Haringey, London", "label": "Haringey", "group": "North London"},
    {"id": "Islington, London", "label": "Islington", "group": "North London"},
    {"id": "Camden, London", "label": "Camden", "group": "North London"},
    {"id": "Finchley, London", "label": "Finchley", "group": "North London"},
    {"id": "Muswell Hill, London", "label": "Muswell Hill", "group": "North London"},
    {"id": "Highgate, London", "label": "Highgate", "group": "North London"},
    {"id": "Crouch End, London", "label": "Crouch End", "group": "North London"},
    {"id": "Tottenham, London", "label": "Tottenham", "group": "North London"},
    {"id": "Wood Green, London", "label": "Wood Green", "group": "North London"},
    {"id": "Archway, London", "label": "Archway", "group": "North London"},
    {"id": "Kentish Town, London", "label": "Kentish Town", "group": "North London"},
    {"id": "Hampstead, London", "label": "Hampstead", "group": "North London"},
    {"id": "Finsbury Park, London", "label": "Finsbury Park", "group": "North London"},
    {"id": "Holloway, London", "label": "Holloway", "group": "North London"},
    {"id": "Angel, London", "label": "Angel", "group": "North London"},
    {"id": "Palmers Green, London", "label": "Palmers Green", "group": "North London"},
    {"id": "Edgware, London", "label": "Edgware", "group": "North London"},
    # Harrow & surrounding areas
    {"id": "Harrow, London", "label": "Harrow", "group": "West London"},
    {"id": "Pinner, London", "label": "Pinner", "group": "West London"},
    {"id": "Stanmore, London", "label": "Stanmore", "group": "West London"},
    {"id": "Northwood, London", "label": "Northwood", "group": "West London"},
    {"id": "Ruislip, London", "label": "Ruislip", "group": "West London"},
    {"id": "Eastcote, London", "label": "Eastcote", "group": "West London"},
    {"id": "Kingsbury, London", "label": "Kingsbury", "group": "West London"},
    {"id": "North London", "label": "North London (General)", "group": "Broader London"},
    {"id": "London", "label": "London (City-wide)", "group": "Broader London"},
]


# ---------------------------------------------------------------------------
# Query filtering — select only queries matching user's filter selections
# ---------------------------------------------------------------------------
def filter_queries(
    business_types: List[str],
    locations: List[str],
) -> List[Tuple[str, str, str, str, str]]:
    """Return only the SEARCH_QUERIES matching the user's selections.

    If ``business_types`` is empty, all business types are included.
    If ``locations`` is empty or contains "all", all locations are included.
    """
    all_bt = not business_types  # empty = select all
    all_loc = not locations or "all" in locations

    filtered = []
    for query_tuple in SEARCH_QUERIES:
        _term, location, _tier, _cat, bt_tag = query_tuple

        bt_match = all_bt or bt_tag in business_types
        loc_match = all_loc or location in locations

        if bt_match and loc_match:
            filtered.append(query_tuple)

    return filtered


# ---------------------------------------------------------------------------
# Borough & area zone inference
# ---------------------------------------------------------------------------
# HIGH-priority boroughs (from master.md)
HIGH_PRIORITY_BOROUGHS: Set[str] = {
    "barnet", "enfield", "haringey", "islington", "camden",
    "finchley", "muswell hill", "highgate", "crouch end",
    "tottenham", "wood green", "archway", "kentish town",
    "hampstead", "finsbury park", "holloway", "angel", "palmers green",
}

# All London boroughs mapped to area zones
BOROUGH_TO_ZONE: Dict[str, str] = {
    # North London
    "barnet": "North London",
    "enfield": "North London",
    "haringey": "North London",
    "islington": "North London",
    "camden": "North London",
    "finchley": "North London",
    "muswell hill": "North London",
    "highgate": "North London",
    "crouch end": "North London",
    "tottenham": "North London",
    "wood green": "North London",
    "archway": "North London",
    "kentish town": "North London",
    "hampstead": "North London",
    "finsbury park": "North London",
    "holloway": "North London",
    "angel": "North London",
    "palmers green": "North London",
    "hackney": "North London",
    "stoke newington": "North London",
    "hornsey": "North London",
    "whetstone": "North London",
    "southgate": "North London",
    "winchmore hill": "North London",
    "edmonton": "North London",
    "totteridge": "North London",

    # Central London
    "westminster": "Central London",
    "city of london": "Central London",
    "covent garden": "Central London",
    "soho": "Central London",
    "mayfair": "Central London",
    "marylebone": "Central London",
    "fitzrovia": "Central London",
    "bloomsbury": "Central London",
    "holborn": "Central London",
    "strand": "Central London",
    "kensington": "Central London",
    "chelsea": "Central London",
    "knightsbridge": "Central London",

    # East London
    "tower hamlets": "East London",
    "newham": "East London",
    "barking": "East London",
    "dagenham": "East London",
    "redbridge": "East London",
    "havering": "East London",
    "waltham forest": "East London",
    "walthamstow": "East London",
    "stratford": "East London",
    "canary wharf": "East London",
    "docklands": "East London",
    "leyton": "East London",
    "leytonstone": "East London",
    "ilford": "East London",
    "romford": "East London",
    "bethnal green": "East London",
    "bow": "East London",
    "mile end": "East London",
    "poplar": "East London",
    "whitechapel": "East London",
    "shoreditch": "East London",

    # West London
    "ealing": "West London",
    "hounslow": "West London",
    "hillingdon": "West London",
    "brent": "West London",
    "harrow": "West London",
    "hammersmith": "West London",
    "fulham": "West London",
    "chiswick": "West London",
    "acton": "West London",
    "shepherds bush": "West London",
    "notting hill": "West London",
    "paddington": "West London",
    "wembley": "West London",
    "uxbridge": "West London",
    "richmond": "West London",
    "twickenham": "West London",
    "pinner": "West London",
    "stanmore": "West London",
    "northwood": "West London",
    "ruislip": "West London",
    "eastcote": "West London",
    "kingsbury": "West London",

    # South London
    "southwark": "South London",
    "lambeth": "South London",
    "lewisham": "South London",
    "greenwich": "South London",
    "bromley": "South London",
    "croydon": "South London",
    "sutton": "South London",
    "merton": "South London",
    "wandsworth": "South London",
    "kingston": "South London",
    "brixton": "South London",
    "peckham": "South London",
    "camberwell": "South London",
    "dulwich": "South London",
    "streatham": "South London",
    "tooting": "South London",
    "wimbledon": "South London",
    "clapham": "South London",
    "battersea": "South London",
    "deptford": "South London",
    "catford": "South London",
    "eltham": "South London",
    "woolwich": "South London",
    "bexley": "South London",
}

# Hertfordshire towns (outside London)
HERTFORDSHIRE_TOWNS: Set[str] = {
    "watford", "st albans", "stevenage", "hemel hempstead", "welwyn",
    "hatfield", "hertford", "potters bar", "borehamwood", "bushey",
    "cheshunt", "hoddesdon", "ware", "bishops stortford", "letchworth",
    "hitchin", "royston", "rickmansworth", "berkhamsted", "tring",
    "harpenden", "radlett", "elstree",
}

# Exclusions — existing clients (case-insensitive match)
EXCLUDED_BUSINESSES: Set[str] = {
    "rendall & rittner",
    "rendall and rittner",
    "mvn block management",
}

# Keywords that trigger HIGH priority when found in business name
HIGH_PRIORITY_NAME_KEYWORDS: List[str] = [
    "block", "leasehold", "rtm", "service charge", "estate management",
]

# Industry mapping from category
INDUSTRY_MAP: Dict[str, str] = {
    CATEGORY_PROPERTY_MGMT: "Property Management",
    CATEGORY_ESTATE_AGENT: "Estate Agency",
}


# ---------------------------------------------------------------------------
# Dataclass for raw business data scraped from the Maps feed
# ---------------------------------------------------------------------------
@dataclass
class RawBusiness:
    """Raw data collected from a Google Maps listing panel."""

    name: str
    website: str
    phone: str
    maps_url: str
    address: str
    search_location: str   # the location string used in the search query
    icp_tier: str          # "Tier 1 – Block Management" or "Tier 2 – Estate Agent"
    category: str          # "Property Management" or "Estate Agent / Letting Agent"
    rating: str = ""
    review_count: int = 0
    notes: str = ""


# ---------------------------------------------------------------------------
# Helper functions
# ---------------------------------------------------------------------------
def _infer_borough(address: str, search_location: str) -> str:
    """Extract London borough from address text or fall back to search location.

    Checks the full address against the known borough/area list.
    Falls back to the search location string (e.g. "Barnet, London" → "Barnet").
    """
    address_lower = address.lower()

    # Try matching known boroughs/areas in the address
    for borough in BOROUGH_TO_ZONE:
        if borough in address_lower:
            return borough.title()

    # Fall back to search location (take the part before the comma)
    if search_location:
        location_parts = search_location.split(",")
        candidate = location_parts[0].strip().lower()
        if candidate in BOROUGH_TO_ZONE or candidate in HERTFORDSHIRE_TOWNS:
            return candidate.title()
        # Check if any known borough is in the search location
        for borough in BOROUGH_TO_ZONE:
            if borough in candidate:
                return borough.title()

    return "Unknown"


def _infer_area_zone(borough: str, address: str) -> str:
    """Map a borough name to its area zone.

    Falls back to checking Hertfordshire towns in the address.
    """
    borough_lower = borough.lower()

    # Direct lookup
    if borough_lower in BOROUGH_TO_ZONE:
        return BOROUGH_TO_ZONE[borough_lower]

    # Check if any Hertfordshire town appears in the address
    address_lower = address.lower()
    for town in HERTFORDSHIRE_TOWNS:
        if town in address_lower:
            return "Hertfordshire"

    # If the address contains "London" at all, classify as North London
    # (since our queries are primarily North London focused)
    if "london" in address_lower:
        return "North London"

    return "Hertfordshire"  # Outside London defaults to Hertfordshire per master.md


def _parse_rating(rating_text: str) -> Tuple[str, int]:
    """Extract numeric rating and review count from Maps aria-label text.

    Example input: "4.5 stars 123 Reviews"
    Returns: ("4.5", 123)
    """
    rating_value = ""
    review_count = 0

    if not rating_text:
        return rating_value, review_count

    # Extract star rating (e.g. "4.5")
    rating_match = re.search(r"([\d.]+)\s*star", rating_text, re.IGNORECASE)
    if rating_match:
        rating_value = rating_match.group(1)

    # Extract review count (e.g. "123")
    review_match = re.search(r"([\d,]+)\s*review", rating_text, re.IGNORECASE)
    if review_match:
        review_count = int(review_match.group(1).replace(",", ""))

    return rating_value, review_count


def _calculate_priority(
    category: str,
    borough: str,
    area_zone: str,
    business_name: str,
    rating_value: str,
    review_count: int,
) -> str:
    """Determine outreach priority (HIGH / MEDIUM / LOW) per master.md rules.

    HIGH →
      - Category is Property Management AND borough is in HIGH_PRIORITY_BOROUGHS
      - OR business name contains: "block", "leasehold", "RTM",
        "service charge", "estate management"

    MEDIUM →
      - Estate/letting agent with rating ≥ 4.0 AND review_count ≥ 20
      - OR Property Management firm outside HIGH boroughs but still in London

    LOW →
      - Estate agent with rating < 4.0 or review_count < 10
      - OR any result outside London (e.g. Hertfordshire)
    """
    name_lower = business_name.lower()
    borough_lower = borough.lower()

    # Check HIGH conditions
    is_property_mgmt = (category == CATEGORY_PROPERTY_MGMT)
    in_high_borough = (borough_lower in HIGH_PRIORITY_BOROUGHS)
    has_high_keyword = any(kw in name_lower for kw in HIGH_PRIORITY_NAME_KEYWORDS)

    if (is_property_mgmt and in_high_borough) or has_high_keyword:
        return "HIGH"

    # Check MEDIUM conditions
    is_estate_agent = (category == CATEGORY_ESTATE_AGENT)
    try:
        numeric_rating = float(rating_value) if rating_value else 0.0
    except (ValueError, TypeError):
        numeric_rating = 0.0

    if is_estate_agent and numeric_rating >= 4.0 and review_count >= 20:
        return "MEDIUM"

    if is_property_mgmt and area_zone != "Hertfordshire":
        # Property management outside HIGH boroughs but still in London
        return "MEDIUM"

    # LOW — everything else
    return "LOW"


def _is_excluded(business_name: str) -> bool:
    """Check whether the business should be excluded (existing client)."""
    return business_name.strip().lower() in EXCLUDED_BUSINESSES


# ---------------------------------------------------------------------------
# AI email enrichment helpers
# ---------------------------------------------------------------------------
# Common patterns for extracting person names near role/title keywords
_NAME_PATTERNS = [
    # "Director: John Smith" / "Manager — Jane Doe"
    re.compile(
        r"(?:director|manager|founder|owner|principal|partner|ceo|md|managing\s+director)"
        r"\s*[:\-–—]\s*([A-Z][a-z]+(?:\s+[A-Z][a-z]+){1,2})",
        re.IGNORECASE,
    ),
    # "John Smith, Director" / "Jane Doe - Founder"
    re.compile(
        r"([A-Z][a-z]+(?:\s+[A-Z][a-z]+){1,2})\s*[,\-–—]\s*"
        r"(?:director|manager|founder|owner|principal|partner|ceo|md|managing\s+director)",
        re.IGNORECASE,
    ),
]


def _extract_person_name(html: str) -> str:
    """Best-effort extraction of a contact person's name from HTML.

    Looks for patterns like "Director: John Smith" or structured team
    sections. Returns the first match or empty string.
    """
    # Remove script/style noise
    soup = BeautifulSoup(html, "html.parser")
    for tag in soup(["script", "style", "noscript"]):
        tag.decompose()
    clean_text = soup.get_text(separator=" ", strip=True)

    for pattern in _NAME_PATTERNS:
        match = pattern.search(clean_text)
        if match:
            name = match.group(1).strip()
            # Basic sanity: name should be 2-4 words, no weird chars
            words = name.split()
            if 2 <= len(words) <= 4 and all(w.isalpha() for w in words):
                return name

    return ""


def _extract_meta_description(html: str) -> str:
    """Extract the meta description from HTML for company description."""
    soup = BeautifulSoup(html, "html.parser")

    # Try <meta name="description">
    meta = soup.find("meta", attrs={"name": "description"})
    if meta and meta.get("content"):
        desc = meta["content"].strip()
        if len(desc) > 20:
            return desc[:500]

    # Try <meta property="og:description">
    og = soup.find("meta", attrs={"property": "og:description"})
    if og and og.get("content"):
        desc = og["content"].strip()
        if len(desc) > 20:
            return desc[:500]

    return ""


# ---------------------------------------------------------------------------
# Excel export
# ---------------------------------------------------------------------------
def export_to_excel(leads: List[Lead], output_dir: str = "output") -> str:
    """Export leads to a colour-coded Excel workbook with 3 sheets.

    Output file: neo_eco_leads_YYYYMMDD_HHMMSS.xlsx

    Sheets:
      1. All Leads — every result, sorted by priority (HIGH first)
      2. High Priority — only HIGH priority rows
      3. By Borough — all results sorted by borough A–Z

    Formatting:
      - Frozen top row (header)
      - Autofilter on all columns
      - Colour-coded outreach_priority cell:
          HIGH   → #C6EFCE (green)
          MEDIUM → #FFEB9C (yellow)
          LOW    → #FFC7CE (red)
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import PatternFill, Font, Alignment
        from openpyxl.utils import get_column_letter
    except ImportError:
        raise RuntimeError(
            "openpyxl is required for Excel export. "
            "Install it with: pip install openpyxl"
        )

    os.makedirs(output_dir, exist_ok=True)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"neo_eco_leads_{timestamp}.xlsx"
    filepath = os.path.join(output_dir, filename)

    # Priority colour fills
    fill_high = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    fill_medium = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
    fill_low = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

    priority_fills = {
        "HIGH": fill_high,
        "MEDIUM": fill_medium,
        "LOW": fill_low,
    }

    # CRM-style export columns (17 primary + internal scoring columns)
    export_headers = [
        "Company Name", "Email", "Person", "Primary_Designation",
        "Company Description", "Country", "Industry", "Website",
        "Employees", "Revenue", "Founded",
        "Company_Phone", "Company_LinkedIn",
        "Primary_Phone", "LinkedIn",
        "Alternate_Person", "Alternate_Designation",
    ]

    # Extra internal columns appended after the CRM columns
    internal_headers = [
        "Outreach_Priority", "ICP_Tier", "Category", "Borough",
        "Area_Zone", "Rating", "Reviews", "Google_Maps_URL", "Address", "Notes",
    ]

    headers = export_headers + internal_headers

    header_font = Font(bold=True, size=11)
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_text_colour = Font(bold=True, size=11, color="FFFFFF")

    # Priority sort order
    priority_order = {"HIGH": 0, "MEDIUM": 1, "LOW": 2}

    def _lead_to_row(lead: Lead) -> list:
        return [
            # CRM columns (17)
            lead.business_name,
            lead.email,
            lead.contact_person,
            lead.primary_designation,
            lead.company_description,
            lead.country,
            lead.industry,
            lead.website,
            lead.employees,
            lead.revenue,
            lead.founded,
            lead.company_phone,
            lead.company_linkedin,
            lead.primary_phone,
            lead.linkedin,
            lead.alternate_person,
            lead.alternate_designation,
            # Internal scoring columns
            lead.outreach_priority,
            lead.icp_tier,
            lead.category,
            lead.borough,
            lead.area_zone,
            lead.rating,
            lead.review_count,
            lead.google_maps_url,
            lead.address,
            lead.notes,
        ]

    def _write_sheet(ws, data: List[Lead], sheet_title: str) -> None:
        """Write headers and data rows to a worksheet."""
        # Write header row
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = header_text_colour
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")

        # Write data rows
        for row_idx, lead in enumerate(data, 2):
            row_values = _lead_to_row(lead)
            for col_idx, value in enumerate(row_values, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)

                # Colour the priority cell
                if col_idx == 1:  # outreach_priority column
                    fill = priority_fills.get(str(value).upper())
                    if fill:
                        cell.fill = fill

        # Freeze top row
        ws.freeze_panes = "A2"

        # Autofilter across all columns
        if data:
            last_col = get_column_letter(len(headers))
            last_row = len(data) + 1
            ws.auto_filter.ref = f"A1:{last_col}{last_row}"

        # Auto-width columns (approximate)
        for col_idx, header in enumerate(headers, 1):
            col_letter = get_column_letter(col_idx)
            max_width = len(header) + 4
            for row_idx in range(2, min(len(data) + 2, 52)):  # sample first 50 rows
                cell_value = str(ws.cell(row=row_idx, column=col_idx).value or "")
                max_width = max(max_width, min(len(cell_value) + 2, 50))
            ws.column_dimensions[col_letter].width = max_width

    wb = Workbook()

    # Sheet 1: All Leads (sorted by priority: HIGH first)
    ws_all = wb.active
    ws_all.title = "All Leads"
    sorted_all = sorted(leads, key=lambda x: priority_order.get(x.outreach_priority, 3))
    _write_sheet(ws_all, sorted_all, "All Leads")

    # Sheet 2: High Priority (only HIGH rows)
    ws_high = wb.create_sheet("High Priority")
    high_only = [lead for lead in leads if lead.outreach_priority == "HIGH"]
    _write_sheet(ws_high, high_only, "High Priority")

    # Sheet 3: By Borough (sorted by borough A–Z)
    ws_borough = wb.create_sheet("By Borough")
    sorted_borough = sorted(leads, key=lambda x: x.borough.lower())
    _write_sheet(ws_borough, sorted_borough, "By Borough")

    wb.save(filepath)
    return filepath


# ---------------------------------------------------------------------------
# CSV export
# ---------------------------------------------------------------------------
def export_to_csv(leads: List[Lead], output_dir: str = "output") -> str:
    """Export leads to a CSV file with all 18 fields.

    Output file: neo_eco_leads_YYYYMMDD_HHMMSS.csv
    """
    os.makedirs(output_dir, exist_ok=True)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"neo_eco_leads_{timestamp}.csv"
    filepath = os.path.join(output_dir, filename)

    # CRM-style export columns (same 17 + internal scoring)
    export_headers = [
        "Company Name", "Email", "Person", "Primary_Designation",
        "Company Description", "Country", "Industry", "Website",
        "Employees", "Revenue", "Founded",
        "Company_Phone", "Company_LinkedIn",
        "Primary_Phone", "LinkedIn",
        "Alternate_Person", "Alternate_Designation",
    ]

    internal_headers = [
        "Outreach_Priority", "ICP_Tier", "Category", "Borough",
        "Area_Zone", "Rating", "Reviews", "Google_Maps_URL", "Address", "Notes",
    ]

    headers = export_headers + internal_headers

    priority_order = {"HIGH": 0, "MEDIUM": 1, "LOW": 2}
    sorted_leads = sorted(leads, key=lambda x: priority_order.get(x.outreach_priority, 3))

    with open(filepath, "w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        writer.writerow(headers)
        for lead in sorted_leads:
            writer.writerow([
                # CRM columns (17)
                lead.business_name, lead.email, lead.contact_person,
                lead.primary_designation, lead.company_description,
                lead.country, lead.industry, lead.website,
                lead.employees, lead.revenue, lead.founded,
                lead.company_phone, lead.company_linkedin,
                lead.primary_phone, lead.linkedin,
                lead.alternate_person, lead.alternate_designation,
                # Internal scoring
                lead.outreach_priority, lead.icp_tier, lead.category,
                lead.borough, lead.area_zone, lead.rating,
                lead.review_count, lead.google_maps_url,
                lead.address, lead.notes,
            ])

    return filepath


# ---------------------------------------------------------------------------
# Terminal summary
# ---------------------------------------------------------------------------
def print_terminal_summary(
    queries_run: int,
    leads: List[Lead],
    output_path: str,
) -> None:
    """Print the exact terminal summary format required by the client."""
    high = sum(1 for lead in leads if lead.outreach_priority == "HIGH")
    medium = sum(1 for lead in leads if lead.outreach_priority == "MEDIUM")
    low = sum(1 for lead in leads if lead.outreach_priority == "LOW")

    print()
    print("=" * 48)
    print("  Neo Eco Cleaning \u2014 Lead Generator Complete")
    print("=" * 48)
    print(f"  Queries run      : {queries_run}")
    print(f"  Total leads      : {len(leads)}")
    print(f"  \U0001f7e2 High Priority : {high}")
    print(f"  \U0001f7e1 Medium        : {medium}")
    print(f"  \U0001f534 Low           : {low}")
    print(f"  Output saved     : {output_path}")
    print("=" * 48)
    print()


# ---------------------------------------------------------------------------
# Main Engine
# ---------------------------------------------------------------------------
class LeadScraperEngine:
    """Orchestrates the full scrape-enrich-score pipeline for Neo Eco Cleaning."""

    RATE_LIMIT_DELAY = 0.8  # seconds between website visits

    def __init__(
        self,
        on_progress: Callable[[ScrapeProgress], None],
        on_lead: Callable[[Lead], None],
    ):
        self.on_progress = on_progress
        self.on_lead = on_lead
        self.leads_found = 0
        self.seen_domains: Set[str] = set()
        self.seen_businesses: Set[str] = set()   # dedup: lowercase(name)
        self.seen_phones: Set[str] = set()       # dedup: normalised phone
        self.all_leads: List[Lead] = []
        self.export_format: str = "xlsx"

    # ---- public entry point ------------------------------------------------
    async def run(self, request: SearchRequest) -> None:
        """Execute all search queries and produce the final lead sheet."""
        self.export_format = request.export_format or "xlsx"

        # Filter queries based on user selections
        queries = filter_queries(request.business_types, request.locations)
        if not queries:
            # No matching queries for the selected filter combination.
            # Do NOT fall back to all queries — respect the user's filter.
            self._emit(
                "completed",
                "No matching queries for your filter selection. Try broadening your filters.",
            )
            return
        total_queries = len(queries)

        self._emit(
            "scraping_maps",
            f"Preparing {total_queries} search queries across London boroughs\u2026",
        )

        async with async_playwright() as pw:
            browser = await pw.chromium.launch(headless=True)

            # --- Pass 1: collect raw business data from Maps ---------------
            all_raw: List[RawBusiness] = []
            maps_ctx = await browser.new_context(
                user_agent=(
                    "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) "
                    "AppleWebKit/537.36 (KHTML, like Gecko) "
                    "Chrome/124.0.0.0 Safari/537.36"
                ),
                viewport={"width": 1280, "height": 900},
                locale="en-GB",
                timezone_id="Europe/London",
                geolocation={"latitude": LONDON_LAT, "longitude": LONDON_LNG},
                permissions=["geolocation"],
            )
            maps_page = await maps_ctx.new_page()

            for idx, (search_term, location, icp_tier, category, _bt) in enumerate(queries, 1):
                query_str = f"{search_term} in {location}"
                self._emit(
                    "scraping_maps",
                    f"[{idx}/{total_queries}] Searching: {query_str}",
                )
                raw = await self._collect_from_maps(
                    maps_page, query_str, location, icp_tier, category,
                )
                all_raw.extend(raw)

            await maps_ctx.close()

            # Deduplicate by domain + business name + phone
            unique_raw: List[RawBusiness] = []
            for biz in all_raw:
                # Exclusion check
                if _is_excluded(biz.name):
                    print(f"[dedup] EXCLUDED (existing client): {biz.name}")
                    continue

                # Domain dedup
                domain = self._domain(biz.website)
                name_key = biz.name.strip().lower()
                phone_key = re.sub(r"[^0-9+]", "", biz.phone) if biz.phone else ""

                is_duplicate = False

                if domain and domain in self.seen_domains:
                    is_duplicate = True
                if name_key and name_key in self.seen_businesses:
                    is_duplicate = True
                if phone_key and phone_key != "" and phone_key in self.seen_phones:
                    is_duplicate = True

                if not is_duplicate:
                    if domain:
                        self.seen_domains.add(domain)
                    if name_key:
                        self.seen_businesses.add(name_key)
                    if phone_key:
                        self.seen_phones.add(phone_key)
                    unique_raw.append(biz)

            self._emit(
                "enriching",
                f"Maps done. {len(all_raw)} total \u2192 {len(unique_raw)} unique businesses. Enriching\u2026",
            )

            # --- Pass 2: enrich each business on its own page --------------
            enrich_ctx = await browser.new_context(
                user_agent=(
                    "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) "
                    "AppleWebKit/537.36 (KHTML, like Gecko) "
                    "Chrome/124.0.0.0 Safari/537.36"
                ),
            )

            for idx, biz in enumerate(unique_raw, 1):
                self._emit(
                    "enriching",
                    f"[{idx}/{len(unique_raw)}] Enriching: {biz.name} ({biz.website})",
                )
                enrich_page = await enrich_ctx.new_page()
                try:
                    lead = await self._enrich(enrich_page, biz)
                    if lead:
                        self.leads_found += 1
                        self.all_leads.append(lead)
                        self.on_lead(lead)
                except Exception as exc:
                    print(f"[enrich] Error on {biz.website}: {exc}")
                finally:
                    await enrich_page.close()
                await asyncio.sleep(self.RATE_LIMIT_DELAY)

            await enrich_ctx.close()
            await browser.close()

        # --- Export & terminal summary ---
        output_path = ""
        if self.all_leads:
            try:
                if self.export_format == "csv":
                    output_path = export_to_csv(self.all_leads)
                else:
                    output_path = export_to_excel(self.all_leads)
            except Exception as exc:
                print(f"[export] Export failed: {exc}")
                output_path = "EXPORT FAILED"

        print_terminal_summary(total_queries, self.all_leads, output_path)

        self._emit(
            "completed",
            f"Done! {self.leads_found} qualified leads found. Saved to {output_path}",
        )

    # ---- Pass 1 helpers ---------------------------------------------------
    async def _collect_from_maps(
        self,
        page: Page,
        query: str,
        search_location: str,
        icp_tier: str,
        category: str,
    ) -> List[RawBusiness]:
        """Open Google Maps centred on London and collect results."""
        results: List[RawBusiness] = []

        # Build a Maps URL with explicit viewport coordinates
        encoded_query = urllib.parse.quote(query)
        url = (
            f"https://www.google.com/maps/search/{encoded_query}"
            f"/@{LONDON_LAT},{LONDON_LNG},{LONDON_ZOOM}z?hl=en"
        )

        try:
            await page.goto(url, timeout=30_000)
            await asyncio.sleep(2)

            # Dismiss cookie consent if present
            for btn_text in ["Accept all", "Reject all", "I agree"]:
                try:
                    btn = page.locator(f"button:has-text('{btn_text}')").first
                    if await btn.is_visible(timeout=2000):
                        await btn.click()
                        await asyncio.sleep(1)
                        break
                except Exception:
                    pass

            # Wait for the results feed
            feed = page.locator('div[role="feed"]')
            try:
                await feed.wait_for(timeout=15_000)
            except Exception:
                print(f"[maps] No feed found for: {query}")
                return results

            # Scroll the feed 3 times to load more results
            for _ in range(3):
                await feed.evaluate("el => el.scrollTop = el.scrollHeight")
                await asyncio.sleep(1.5)

            # Collect every business card link
            links = await page.locator('a[href*="/maps/place/"]').all()
            self._emit("scraping_maps", f"Found {len(links)} listings for: {query}")

            for link in links:
                try:
                    name = await link.get_attribute("aria-label", timeout=3000)
                    href = await link.get_attribute("href", timeout=3000)
                    if not name or not href:
                        continue

                    # Click the listing to open the side panel
                    await link.click(timeout=5000)
                    await asyncio.sleep(1)

                    # Extract website
                    website = await self._extract_panel_field(
                        page, 'a[data-item-id="authority"]', "href"
                    )
                    if not website:
                        website = ""

                    # Extract phone
                    phone = await self._extract_panel_field(
                        page,
                        'button[data-item-id^="phone:tel:"]',
                        "aria-label",
                    )
                    if phone:
                        phone = (
                            phone.replace("Phone: ", "")
                            .replace("Phone number: ", "")
                            .strip()
                        )
                    else:
                        phone = ""

                    # Extract full address from panel
                    address = await self._extract_panel_field(
                        page,
                        'button[data-item-id="address"]',
                        "aria-label",
                    )
                    if address:
                        address = address.replace("Address: ", "").strip()
                    else:
                        address = ""

                    # Location verification — must be London or surrounding area
                    address_lower = address.lower()
                    maps_url_lower = (href or "").lower()
                    location_match = any(
                        kw in address_lower or kw in maps_url_lower
                        for kw in [
                            "london", "barnet", "enfield", "haringey",
                            "islington", "camden", "north london",
                            "n1", "n2", "n3", "n4", "n5", "n6", "n7",
                            "n8", "n9", "n10", "n11", "n12", "n13",
                            "n14", "n15", "n16", "n17", "n18", "n19", "n20",
                            "n22", "nw1", "nw2", "nw3", "nw5", "nw6",
                            "ec1", "ec2", "wc1", "wc2",
                            "hertfordshire", "herts",
                            "uk", "england", "united kingdom",
                        ]
                    )

                    if not location_match and address:
                        print(f"[maps] SKIPPED (wrong location): {name} \u2014 Address: {address}")
                        continue

                    # Extract rating
                    rating_text = ""
                    try:
                        rating_el = page.locator('div[role="img"][aria-label*="stars"]').first
                        rating_text = (await rating_el.get_attribute("aria-label", timeout=2000)) or ""
                    except Exception:
                        pass

                    rating_value, review_count = _parse_rating(rating_text)

                    # Extract notes / description from the Maps panel
                    notes = ""
                    try:
                        about_selectors = [
                            'div.WeS02d.fontBodyMedium',
                            'div[class*="PYvSYb"]',
                            'div.rogA2c div.fontBodyMedium',
                        ]
                        for sel in about_selectors:
                            try:
                                about_el = page.locator(sel).first
                                txt = (await about_el.text_content(timeout=2000)) or ""
                                txt = txt.strip()
                                if txt and len(txt) > 15:
                                    notes = txt[:300]
                                    break
                            except Exception:
                                continue
                    except Exception:
                        pass

                    results.append(
                        RawBusiness(
                            name=name.strip(),
                            website=website.strip() if website else "",
                            phone=phone,
                            maps_url=href,
                            address=address,
                            search_location=search_location,
                            icp_tier=icp_tier,
                            category=category,
                            rating=rating_value,
                            review_count=review_count,
                            notes=notes,
                        )
                    )

                except Exception:
                    continue

        except Exception as exc:
            print(f"[maps] Fatal error for query '{query}': {exc}")

        return results

    @staticmethod
    async def _extract_panel_field(
        page: Page, selector: str, attribute: str
    ) -> Optional[str]:
        try:
            els = await page.locator(selector).all()
            if els:
                return await els[0].get_attribute(attribute, timeout=3000)
        except Exception:
            pass
        return None

    # ---- Pass 2: website enrichment ---------------------------------------
    async def _enrich(
        self, page: Page, biz: RawBusiness,
    ) -> Optional[Lead]:
        """Visit the company website to extract email, person name,
        and company description, then score the lead."""
        emails: Set[str] = set()
        contact_person = ""
        company_description = ""

        # ---- Visit homepage (if present) ----
        if biz.website:
            try:
                await page.goto(biz.website, timeout=12_000, wait_until="domcontentloaded")
                await asyncio.sleep(1)
                html = await page.content()

                # Extract emails from homepage HTML
                self._extract_emails(html, emails)

                # Extract meta description for company_description
                company_description = _extract_meta_description(html)

                # Try to extract person name from homepage
                contact_person = _extract_person_name(html)

                # If no emails or person found on homepage, try sub-pages
                if not emails or not contact_person:
                    soup = BeautifulSoup(html, "html.parser")
                    internal_links = soup.find_all("a", href=True)
                    contact_pages: List[str] = []
                    for a in internal_links:
                        href_lower = a["href"].lower()
                        if any(
                            kw in href_lower
                            for kw in ["contact", "about", "team", "people"]
                        ):
                            full = (
                                a["href"]
                                if a["href"].startswith("http")
                                else urllib.parse.urljoin(biz.website, a["href"])
                            )
                            contact_pages.append(full)

                    # Visit up to 2 sub-pages
                    for sub_url in list(set(contact_pages))[:2]:
                        try:
                            await page.goto(
                                sub_url, timeout=10_000, wait_until="domcontentloaded"
                            )
                            await asyncio.sleep(0.5)
                            sub_html = await page.content()
                            self._extract_emails(sub_html, emails)

                            # Try person name extraction on sub-pages
                            if not contact_person:
                                contact_person = _extract_person_name(sub_html)

                            # Fallback company description from sub-pages
                            if not company_description:
                                company_description = _extract_meta_description(sub_html)
                        except Exception:
                            pass

            except Exception as exc:
                print(f"[enrich] Could not load {biz.website}: {exc}")

        # Clean emails — remove junk domains
        clean_emails = sorted(
            {
                e
                for e in emails
                if not any(
                    junk in e
                    for junk in [
                        "sentry", "wix", "example", ".png", ".jpg",
                        ".jpeg", ".gif", ".svg", "cloudflare", "schema.org",
                        "sentry.io", "googleapis", "w3.org",
                    ]
                )
            }
        )
        email_str = ", ".join(clean_emails) if clean_emails else ""

        # Fallback company description to Maps notes
        if not company_description and biz.notes:
            company_description = biz.notes

        # ---- Borough & area zone inference ----
        borough = _infer_borough(biz.address, biz.search_location)
        area_zone = _infer_area_zone(borough, biz.address)

        # ---- Industry from category ----
        industry = INDUSTRY_MAP.get(biz.category, "Property Services")

        # ---- Priority scoring ----
        priority = _calculate_priority(
            category=biz.category,
            borough=borough,
            area_zone=area_zone,
            business_name=biz.name,
            rating_value=biz.rating,
            review_count=biz.review_count,
        )

        return Lead(
            business_name=biz.name,
            email=email_str,
            contact_person=contact_person,
            primary_designation="",             # best-effort; not yet extractable
            company_description=company_description,
            country="United Kingdom",
            industry=industry,
            website=biz.website if biz.website else "",
            employees="",                        # not available from Maps
            revenue="",                          # not available from Maps
            founded="",                          # not available from Maps
            company_phone=biz.phone if biz.phone else "",
            company_linkedin="",                 # not available from Maps
            primary_phone=biz.phone if biz.phone else "",  # same as company phone
            linkedin="",                         # not available from Maps
            alternate_person="",
            alternate_designation="",
            # Internal scoring fields
            outreach_priority=priority,
            icp_tier=biz.icp_tier,
            category=biz.category,
            borough=borough,
            area_zone=area_zone,
            google_maps_url=biz.maps_url,
            rating=biz.rating if biz.rating else "",
            review_count=biz.review_count,
            address=biz.address if biz.address else "",
            notes=biz.notes if biz.notes else "",
        )

    # ---- utilities --------------------------------------------------------
    @staticmethod
    def _extract_emails(html: str, output: Set[str]) -> None:
        """Find all email addresses in raw HTML."""
        pattern = r"[a-zA-Z0-9._%+\-]+@[a-zA-Z0-9.\-]+\.[a-zA-Z]{2,}"
        for addr in re.findall(pattern, html):
            output.add(addr.lower())

    @staticmethod
    def _domain(url: str) -> Optional[str]:
        try:
            return urllib.parse.urlparse(url).netloc.replace("www.", "")
        except Exception:
            return None

    def _emit(self, status: str, action: str) -> None:
        self.on_progress(
            ScrapeProgress(
                status=status,
                leads_found=self.leads_found,
                current_action=action,
            )
        )
